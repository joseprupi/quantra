"""BoE SONIA OIS curve connector tests (hermetic).

Builds a tiny in-memory XLSX workbook that mimics the real BoE OIS
``"4. spot curve"`` sheet (a ``years:`` tenor header on row 4 giving
continuously-compounded zero rates, daily rows from row 6), zips it like
the published feed, and feeds it to :class:`BoeOisCurveAdapter` with the
network download stubbed out. Asserts the connector converts zeros ->
discount factors -> par swap rates (the strictly-correct construction),
not the raw zero. No network access — the real feed is only exercised by
the live proof.
"""

from __future__ import annotations

import io
import math
import zipfile
from datetime import UTC, date, datetime

from openpyxl import Workbook

from quantra_md_ingester.connectors.boe_ois_curve import BoeOisCurveAdapter
from quantra_md_ingester.connectors.boe_ois_curve import (
    BoeOisCurveSeriesSpec as Spec,
)

# Full 0.5-year zero grid (percent), like the real workbook (0.5..N). Upward
# curve so par != zero by a visible margin.
_GRID: dict[float, float] = {
    0.5: 3.80,
    1.0: 3.90,
    1.5: 3.95,
    2.0: 4.00,
    2.5: 4.05,
    3.0: 4.10,
}
_ROWS: tuple[tuple[date, dict[float, float]], ...] = (
    (date(2026, 7, 13), {k: v - 0.05 for k, v in _GRID.items()}),
    (date(2026, 7, 14), _GRID),
)


def _build_fixture_zip() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "4. spot curve"
    ws["A1"] = "UK OIS spot curve"
    ws["A3"] = "Maturity"
    tenors = sorted(_GRID)
    ws.cell(row=4, column=1, value="years:")
    for offset, tenor in enumerate(tenors, start=2):
        ws.cell(row=4, column=offset, value=tenor)
    for row_offset, (day, grid) in enumerate(_ROWS, start=6):
        ws.cell(row=row_offset, column=1, value=datetime(day.year, day.month, day.day))
        for col_offset, tenor in enumerate(tenors, start=2):
            ws.cell(row=row_offset, column=col_offset, value=grid[tenor])
    wb.create_sheet("2. fwd curve")  # a decoy sheet the adapter must ignore

    xbuf = io.BytesIO()
    wb.save(xbuf)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w") as zf:
        zf.writestr("OIS daily data current month.xlsx", xbuf.getvalue())
    return zbuf.getvalue()


class _StubbedAdapter(BoeOisCurveAdapter):
    def __init__(self, payload: bytes) -> None:
        super().__init__()
        self._payload = payload
        self.requested_urls: list[str] = []

    def _download_zip(self, url: str) -> bytes:
        self.requested_urls.append(url)
        return self._payload


def _spec(years: float, tenor: str) -> Spec:
    return Spec(
        maturity_years=years,
        canonical_id=f"GBP.RATES.BOE.OIS.{tenor}.PAR",
        tenor=tenor,
        description=tenor,
    )


def _df(zeros: dict[float, float], t: float) -> float:
    return math.exp(-(zeros[t] / 100.0) * t)


def test_emits_par_rate_not_raw_zero() -> None:
    adapter = _StubbedAdapter(_build_fixture_zip())
    records = adapter.fetch_series(
        [_spec(0.5, "6M"), _spec(1.0, "1Y"), _spec(2.0, "2Y"), _spec(3.0, "3Y")],
        as_of=date(2026, 7, 14),
        start_date=date(2026, 7, 14),
    )
    assert len(records) == 4
    by_id = {r.canonical_id: r for r in records}

    # 6M single-period par (simple) rate from DF(0.5).
    d6m = _df(_GRID, 0.5)
    exp_6m = (1.0 - d6m) / (0.5 * d6m)
    assert by_id["GBP.RATES.BOE.OIS.6M.PAR"].value == exp_6m

    # 3Y annual par swap rate from DF(1..3); must differ from the raw zero.
    ann = sum(_df(_GRID, y) for y in (1.0, 2.0, 3.0))
    exp_3y = (1.0 - _df(_GRID, 3.0)) / ann
    rec_3y = by_id["GBP.RATES.BOE.OIS.3Y.PAR"]
    assert math.isclose(rec_3y.value, exp_3y, rel_tol=0, abs_tol=1e-12)
    # The correction is real: par differs from the (upward) zero here.
    raw_zero_3y = _GRID[3.0] / 100.0
    assert rec_3y.value != raw_zero_3y
    assert abs(rec_3y.value - raw_zero_3y) > 1e-4  # > 1bp

    # Provenance meta records the derivation + the raw zero it came from.
    assert rec_3y.meta["quote_kind"] == "par_swap_rate"
    assert rec_3y.meta["source_sheet"] == "4. spot curve"
    assert rec_3y.meta["raw_zero_rate"] == raw_zero_3y
    assert rec_3y.as_of == datetime(2026, 7, 14, tzinfo=UTC)


def test_bootstrap_recovers_zero_curve() -> None:
    # Feeding the emitted par rates into an annual single-curve bootstrap must
    # recover the original discount factors. Emulate the bootstrap by solving
    # forward: DF(N) = (1 - S_N * sum_{i<N} DF(i)) / (1 + S_N).
    adapter = _StubbedAdapter(_build_fixture_zip())
    specs = [_spec(float(n), f"{n}Y") for n in (1, 2, 3)]
    records = adapter.fetch_series(specs, as_of=date(2026, 7, 14), start_date=date(2026, 7, 14))
    par = {r.meta["maturity_years"]: r.value for r in records}

    dfs: dict[int, float] = {}
    for n in (1, 2, 3):
        s = par[float(n)]
        prior = sum(dfs[i] for i in range(1, n))
        dfs[n] = (1.0 - s * prior) / (1.0 + s)
        assert math.isclose(dfs[n], _df(_GRID, float(n)), rel_tol=0, abs_tol=1e-12)


def test_latest_zip_used_for_current_month_window() -> None:
    adapter = _StubbedAdapter(_build_fixture_zip())
    adapter.fetch_series([_spec(1.0, "1Y")], as_of=date(2026, 7, 14), start_date=date(2026, 7, 1))
    assert adapter.requested_urls == [
        "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/"
        "latest-yield-curve-data.zip"
    ]


def test_archive_zip_used_when_window_reaches_prior_month() -> None:
    adapter = _StubbedAdapter(_build_fixture_zip())
    adapter.fetch_series([_spec(1.0, "1Y")], as_of=date(2026, 7, 14), start_date=date(2026, 6, 1))
    assert adapter.requested_urls == [
        "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/oisddata.zip"
    ]


def test_date_window_is_honoured() -> None:
    adapter = _StubbedAdapter(_build_fixture_zip())
    records = adapter.fetch_series(
        [_spec(1.0, "1Y"), _spec(2.0, "2Y")],
        as_of=date(2026, 7, 14),
        start_date=date(2026, 7, 13),
    )
    assert {r.as_of.date() for r in records} == {date(2026, 7, 13), date(2026, 7, 14)}
    assert len(records) == 4


def test_tenor_not_in_workbook_is_skipped() -> None:
    adapter = _StubbedAdapter(_build_fixture_zip())
    # 5Y depends on zeros at 4Y/5Y which are absent from the fixture grid.
    records = adapter.fetch_series(
        [_spec(5.0, "5Y"), _spec(2.0, "2Y")],
        as_of=date(2026, 7, 14),
        start_date=date(2026, 7, 14),
    )
    assert {r.canonical_id for r in records} == {"GBP.RATES.BOE.OIS.2Y.PAR"}


def test_empty_specs_returns_empty_without_download() -> None:
    class _Boom(BoeOisCurveAdapter):
        def _download_zip(self, url: str) -> bytes:
            raise AssertionError("must not download for empty specs")

    assert _Boom().fetch_series([], as_of=date(2026, 7, 14)) == []
