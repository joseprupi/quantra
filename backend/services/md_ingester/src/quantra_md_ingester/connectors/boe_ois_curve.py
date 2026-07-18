"""BoE SONIA OIS curve connector — strictly-correct par rates from zeros.

The Bank of England publishes a genuine sterling **OIS (SONIA) curve** in
its yield-curve series. The workbooks expose (among others) a
``"4. spot curve"`` sheet: a ``years:`` tenor header on row 4 and one
daily row per business day, giving the **zero-coupon (spot) rates** of the
fitted OIS curve at 0.5-year steps out to 25Y.

**Correctness (par vs zero).** The OIS workbook carries *no* par-yield
sheet — only forward and spot (zero) sheets. Feeding the published zero
rates straight into a par ``SwapHelper`` would misprice by the zero-vs-par
spread (empirically up to ~8bp on the current GBP curve). Instead this
connector builds the curve *correctly from the zeros*:

* BoE spot rates are **continuously compounded** (verified: integrating the
  workbook's instantaneous-forward sheet reproduces the spot sheet to
  <0.5bp), so ``DF(t) = exp(-z(t) * t)`` with ``t`` the maturity in years.
* For each configured pillar it derives the **par OIS swap rate** implied by
  those discount factors — a single-curve (discount = forwarding) OIS with
  an annual fixed leg::

      S_N = (1 - DF(N)) / sum_{i=1..N} DF(i)          (integer-year pillars)
      S_6M = (1 - DF(0.5)) / (0.5 * DF(0.5))          (6M short pillar)

  This is a lossless, deterministic re-expression of the same curve at the
  pillar tenors; feeding these par rates into an annual ``SwapHelper`` (or a
  6M ``DepositHelper``) reconstructs the BoE OIS discount factors exactly.

The emitted canonical ids use the ``PAR`` field
(``GBP.RATES.BOE.OIS.{tenor}.PAR``) so the series is honestly labelled as a
derived par rate, not the raw zero. The full derivation (source sheet,
compounding, method, raw zero) is recorded in ``meta``.

**Freshness.** Like the nominal connector, this distinguishes an *archive*
ZIP (``oisddata.zip`` — full history, but lags ~2 weeks) from the *latest*
ZIP (``latest-yield-curve-data.zip`` — current month, carries yesterday's
fixing). A request whose window starts in the current month pulls the
fresher latest ZIP; anything reaching into a prior month pulls the archive.
"""

from __future__ import annotations

import io
import math
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from typing import Final
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from quantra_common.logging import get_logger
from quantra_md_ingester.models import QuoteRecord

logger = get_logger(__name__)

BOE_OIS_ARCHIVE_ZIP: Final[str] = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/oisddata.zip"
)
BOE_OIS_LATEST_ZIP: Final[str] = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/"
    "latest-yield-curve-data.zip"
)
SPOT_CURVE_SHEET: Final[str] = "4. spot curve"
_TENOR_HEADER_ROW: Final[int] = 4
_DATA_FIRST_ROW: Final[int] = 6
# Tolerance for treating a maturity-in-years grid value as an integer year.
_INT_TOL: Final[float] = 1e-6


@dataclass(frozen=True, slots=True)
class BoeOisCurveSeriesSpec:
    maturity_years: float
    canonical_id: str
    tenor: str
    description: str


class BoeOisCurveAdapter:
    """Download the BoE OIS curve ZIP and emit par OIS swap rates per tenor."""

    def __init__(self, timeout_seconds: int = 60) -> None:
        self.timeout_seconds = timeout_seconds

    def fetch_series(
        self,
        specs: Sequence[BoeOisCurveSeriesSpec],
        as_of: date,
        start_date: date | None = None,
    ) -> list[QuoteRecord]:
        if len(specs) == 0:
            return []

        start = start_date or as_of
        # The "latest" ZIP only contains the current month; older ranges need
        # the archive ZIP (which lags ~2 weeks). Mirrors the nominal connector.
        use_archive = start < as_of.replace(day=1)
        url = BOE_OIS_ARCHIVE_ZIP if use_archive else BOE_OIS_LATEST_ZIP

        payload = self._download_zip(url)
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            candidate_files = [
                n for n in zf.namelist() if n.lower().endswith(".xlsx") and "ois" in n.lower()
            ]
            if len(candidate_files) == 0:
                msg = "No OIS curve workbook found in BoE ZIP payload"
                raise RuntimeError(msg)

            out: list[QuoteRecord] = []
            for workbook_name in candidate_files:
                out.extend(
                    self._parse_workbook(zf.read(workbook_name), workbook_name, specs, start, as_of)
                )
        return out

    def _download_zip(self, url: str) -> bytes:
        request = Request(  # noqa: S310 - constant HTTPS host
            url,
            headers={"User-Agent": "Mozilla/5.0 (compatible; QuantraMarketDataBot/1.0)"},
        )
        with urlopen(request, timeout=self.timeout_seconds) as response:  # noqa: S310
            data: bytes = response.read()
            return data

    def _parse_workbook(
        self,
        workbook_bytes: bytes,
        workbook_name: str,
        specs: Sequence[BoeOisCurveSeriesSpec],
        start_date: date,
        end_date: date,
    ) -> list[QuoteRecord]:
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
        if SPOT_CURVE_SHEET not in wb.sheetnames:
            return []

        ws = wb[SPOT_CURVE_SHEET]
        tenors_row = next(
            ws.iter_rows(min_row=_TENOR_HEADER_ROW, max_row=_TENOR_HEADER_ROW, values_only=True)
        )
        # Full 0.5-year zero grid: maturity (years) -> workbook column index.
        year_to_col: dict[float, int] = {}
        for idx, raw in enumerate(tenors_row):
            if idx == 0 or raw is None:
                continue
            if isinstance(raw, int | float):
                year_to_col[float(raw)] = idx

        out: list[QuoteRecord] = []
        for row in ws.iter_rows(min_row=_DATA_FIRST_ROW, values_only=True):
            raw_date = row[0]
            if not isinstance(raw_date, datetime):
                continue
            as_of_date = raw_date.date()
            if as_of_date < start_date or as_of_date > end_date:
                continue

            zeros = self._zero_grid(row, year_to_col)
            if not zeros:
                continue
            for spec in specs:
                converted = self._par_rate_for_spec(spec, zeros)
                if converted is None:
                    continue
                par_rate, raw_zero = converted
                out.append(
                    QuoteRecord(
                        canonical_id=spec.canonical_id,
                        as_of=datetime.combine(as_of_date, datetime.min.time(), tzinfo=UTC),
                        value=par_rate,
                        source="BOE",
                        vendor_id=f"OIS_PAR_{spec.maturity_years:g}Y",
                        quality_flags={},
                        meta={
                            "dataset": "OIS_PAR_FROM_SPOT_CURVE",
                            "workbook": workbook_name,
                            "maturity_years": spec.maturity_years,
                            "tenor": spec.tenor,
                            "vendor_unit": "percent",
                            "normalized_unit": "decimal_rate",
                            "quote_kind": "par_swap_rate",
                            "source_sheet": SPOT_CURVE_SHEET,
                            "derivation": (
                                "par OIS swap rate (annual fixed, single curve) from "
                                "DF(t)=exp(-z(t)*t) on the BoE continuously-compounded "
                                "spot (zero) curve"
                            ),
                            "raw_zero_rate": raw_zero,
                            "series_description": spec.description,
                        },
                    )
                )
        return out

    @staticmethod
    def _zero_grid(row: tuple[object, ...], year_to_col: dict[float, int]) -> dict[float, float]:
        """Return {maturity_years: zero_rate_decimal} for the populated grid."""

        zeros: dict[float, float] = {}
        for years, col in year_to_col.items():
            raw = row[col]
            if isinstance(raw, int | float):
                zeros[years] = float(raw) / 100.0
        return zeros

    @classmethod
    def _par_rate_for_spec(
        cls,
        spec: BoeOisCurveSeriesSpec,
        zeros: dict[float, float],
    ) -> tuple[float, float] | None:
        """Convert the zero curve to the par rate for ``spec``'s pillar.

        Returns ``(par_rate, raw_zero_at_pillar)`` or ``None`` if the pillar
        (or an integer-year node it depends on) is missing from the grid.
        """

        maturity = spec.maturity_years
        raw_zero = zeros.get(maturity)
        if raw_zero is None:
            return None

        def df(t: float) -> float:
            return math.exp(-zeros[t] * t)

        # 6M short pillar: single-period par (simple) rate.
        if abs(maturity - 0.5) < _INT_TOL:
            d = df(0.5)
            return (1.0 - d) / (0.5 * d), raw_zero

        # Integer-year pillars: annual fixed-leg par swap rate. Needs a zero
        # at every integer year up to N (all present in the 0.5-step grid).
        n = round(maturity)
        if abs(maturity - n) > _INT_TOL:
            # Non-integer, non-6M tenor: no annual schedule — skip rather than
            # emit a wrong number.
            return None
        years = list(range(1, n + 1))
        if any(float(y) not in zeros for y in years):
            return None
        annuity = sum(df(float(y)) for y in years)
        if annuity <= 0.0:
            return None
        par = (1.0 - df(float(n))) / annuity
        return par, raw_zero
