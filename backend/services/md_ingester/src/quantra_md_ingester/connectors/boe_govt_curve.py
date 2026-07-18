"""BoE government nominal yield curve workbook connector.

Pulls the published ZIP of monthly XLSX workbooks, picks the
``"4. spot curve"`` sheet, and emits one ``QuoteRecord`` per
(maturity, business day) tuple. ``openpyxl`` is the only non-stdlib
dependency this connector pulls in; it is added to the service's
``pyproject.toml`` (not to ``quantra-common``) because no other
service consumes XLSX feeds.
"""

from __future__ import annotations

import io
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from typing import Final
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from quantra_common.logging import get_logger
from quantra_md_ingester.models import QuoteRecord

logger = get_logger(__name__)

BOE_GLC_NOMINAL_DAILY_ZIP: Final[str] = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/glcnominalddata.zip"
)
BOE_GLC_NOMINAL_LATEST_ZIP: Final[str] = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/"
    "latest-yield-curve-data.zip"
)
SPOT_CURVE_SHEET: Final[str] = "4. spot curve"
_TENOR_HEADER_ROW: Final[int] = 4
_DATA_FIRST_ROW: Final[int] = 6


@dataclass(frozen=True, slots=True)
class BoeGovtCurveSeriesSpec:
    maturity_years: float
    canonical_id: str
    tenor: str
    description: str


class BoeGovtCurveAdapter:
    """Download + parse the BoE nominal-curve ZIP into per-tenor records."""

    def __init__(self, timeout_seconds: int = 60) -> None:
        self.timeout_seconds = timeout_seconds

    def fetch_series(
        self,
        specs: Sequence[BoeGovtCurveSeriesSpec],
        as_of: date,
        start_date: date | None = None,
    ) -> list[QuoteRecord]:
        if len(specs) == 0:
            return []

        start = start_date or as_of
        # The "latest" ZIP only contains the current month; older ranges
        # need the archive ZIP. The threshold matches what the legacy code
        # did: anything before the first of the current month is treated
        # as historical and pulled from the archive.
        use_archive = start < as_of.replace(day=1)
        url = BOE_GLC_NOMINAL_DAILY_ZIP if use_archive else BOE_GLC_NOMINAL_LATEST_ZIP

        payload = self._download_zip(url)
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            candidate_files = [
                n for n in zf.namelist() if n.lower().endswith(".xlsx") and "nominal" in n.lower()
            ]
            if len(candidate_files) == 0:
                msg = "No nominal curve workbook found in BoE ZIP payload"
                raise RuntimeError(msg)

            out: list[QuoteRecord] = []
            for workbook_name in candidate_files:
                out.extend(
                    self._parse_workbook(zf.read(workbook_name), workbook_name, specs, start, as_of)
                )
        return out

    def _download_zip(self, url: str) -> bytes:
        request = Request(  # noqa: S310 - constant HTTPS host
            url,
            headers={"User-Agent": "Mozilla/5.0 (compatible; QuantraMarketDataBot/1.0)"},
        )
        with urlopen(request, timeout=self.timeout_seconds) as response:  # noqa: S310
            data: bytes = response.read()
            return data

    def _parse_workbook(
        self,
        workbook_bytes: bytes,
        workbook_name: str,
        specs: Sequence[BoeGovtCurveSeriesSpec],
        start_date: date,
        end_date: date,
    ) -> list[QuoteRecord]:
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
        if SPOT_CURVE_SHEET not in wb.sheetnames:
            return []

        ws = wb[SPOT_CURVE_SHEET]
        tenors_row = next(
            ws.iter_rows(min_row=_TENOR_HEADER_ROW, max_row=_TENOR_HEADER_ROW, values_only=True)
        )
        year_to_col: dict[float, int] = {}
        for idx, raw in enumerate(tenors_row):
            if idx == 0 or raw is None:
                continue
            if isinstance(raw, int | float):
                year_to_col[float(raw)] = idx

        active_specs = [s for s in specs if s.maturity_years in year_to_col]
        if len(active_specs) == 0:
            return []

        out: list[QuoteRecord] = []
        for row in ws.iter_rows(min_row=_DATA_FIRST_ROW, values_only=True):
            raw_date = row[0]
            if not isinstance(raw_date, datetime):
                continue
            as_of_date = raw_date.date()
            if as_of_date < start_date or as_of_date > end_date:
                continue

            for spec in active_specs:
                col = year_to_col[spec.maturity_years]
                raw_percent = row[col]
                if raw_percent is None or not isinstance(raw_percent, int | float):
                    continue
                out.append(
                    QuoteRecord(
                        canonical_id=spec.canonical_id,
                        as_of=datetime.combine(as_of_date, datetime.min.time(), tzinfo=UTC),
                        value=float(raw_percent) / 100.0,
                        source="BOE",
                        vendor_id=f"GLCNOMINAL_SPOT_{spec.maturity_years:g}Y",
                        quality_flags={},
                        meta={
                            "dataset": "GLC_NOMINAL_SPOT_CURVE",
                            "workbook": workbook_name,
                            "maturity_years": spec.maturity_years,
                            "tenor": spec.tenor,
                            "vendor_unit": "percent",
                            "normalized_unit": "decimal_rate",
                            "raw_value_percent": float(raw_percent),
                            "series_description": spec.description,
                        },
                    )
                )
        return out
